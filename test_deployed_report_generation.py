"""
Test the deployed backend report generation end-to-end
"""
import requests
import json
from datetime import datetime

BACKEND_URL = "https://sales-and-operation-planning-platform-1.onrender.com/api/v1"
FRONTEND_URL = "https://soptest.netlify.app"

print("=" * 80)
print("TESTING DEPLOYED BACKEND - FULL REPORT GENERATION FLOW")
print("=" * 80)
print(f"Backend: {BACKEND_URL}")
print(f"Frontend: {FRONTEND_URL}")
print("=" * 80)

# Step 1: Test backend health
print("\n[STEP 1] Testing backend health...")
try:
    health = requests.get(f"{BACKEND_URL.replace('/api/v1', '')}/", timeout=10)
    print(f"Status: {health.status_code}")
    print(f"Response: {health.json()}")
except Exception as e:
    print(f"[ERROR] Backend health check failed: {e}")
    exit(1)

# Step 2: Login
print("\n[STEP 2] Logging in...")
try:
    login_response = requests.post(f"{BACKEND_URL}/auth/login", json={
        "email": "admin@sopportal.com",
        "password": "admin123"
    }, timeout=30)

    if login_response.status_code == 200:
        token = login_response.json().get("access_token")
        print(f"[OK] Login successful!")
        print(f"Token: {token[:30]}...")
    else:
        print(f"[ERROR] Login failed: {login_response.status_code}")
        print(login_response.text)
        exit(1)
except Exception as e:
    print(f"[ERROR] Login request failed: {e}")
    exit(1)

headers = {
    "Authorization": f"Bearer {token}",
    "Content-Type": "application/json"
}

# Step 3: Check cycles endpoint (console shows error here)
print("\n[STEP 3] Testing cycles endpoint (console shows error)...")
try:
    cycles_response = requests.get(f"{BACKEND_URL}/cycles", headers=headers, timeout=30)
    print(f"Status: {cycles_response.status_code}")
    if cycles_response.status_code == 200:
        cycles_data = cycles_response.json()
        print(f"Response: {json.dumps(cycles_data, indent=2)}")
    else:
        print(f"[ERROR] Cycles endpoint failed: {cycles_response.text}")
except Exception as e:
    print(f"[ERROR] Cycles request failed: {e}")

# Step 4: Check if November data exists
print("\n[STEP 4] Checking November 2025 data via backend...")
# We'll generate a report to see if data exists

# Step 5: Generate report WITHOUT filters (all data)
print("\n[STEP 5] Generating report WITHOUT date filters (all data)...")
try:
    report_params = {
        "reportType": "sales_summary",
        "format": "excel",
        "includeCharts": True,
        "includeRawData": False
    }

    print(f"Request params: {json.dumps(report_params, indent=2)}")

    report_response = requests.post(
        f"{BACKEND_URL}/reports/generate-instant",
        headers=headers,
        json=report_params,
        timeout=60
    )

    if report_response.status_code == 200:
        print(f"[OK] Report generated! Size: {len(report_response.content)} bytes")

        filename = f"test_all_data_{datetime.now().strftime('%H%M%S')}.xlsx"
        with open(filename, 'wb') as f:
            f.write(report_response.content)
        print(f"[OK] Saved as: {filename}")

        # Try to read the Excel file
        try:
            import openpyxl
            wb = openpyxl.load_workbook(filename)
            sheet = wb[wb.sheetnames[0]]

            print("\n[CHECKING REPORT CONTENTS]")
            for row_idx in range(1, min(25, sheet.max_row + 1)):
                row = []
                for col_idx in range(1, min(3, sheet.max_column + 1)):
                    cell = sheet.cell(row_idx, col_idx)
                    if cell.value:
                        row.append(str(cell.value))
                if row:
                    print(f"  Row {row_idx}: {' | '.join(row)}")

            wb.close()
        except ImportError:
            print("[INFO] Install openpyxl to read Excel contents: pip install openpyxl")
        except Exception as e:
            print(f"[WARNING] Could not read Excel: {e}")

    else:
        print(f"[ERROR] Report generation failed: {report_response.status_code}")
        print(f"Response: {report_response.text}")
except Exception as e:
    print(f"[ERROR] Report request failed: {e}")

# Step 6: Generate report WITH November filter
print("\n[STEP 6] Generating report WITH November 2025 filter...")
try:
    report_params_nov = {
        "reportType": "sales_summary",
        "format": "excel",
        "year": 2025,
        "month": 11,
        "includeCharts": True,
        "includeRawData": False
    }

    print(f"Request params: {json.dumps(report_params_nov, indent=2)}")

    report_response_nov = requests.post(
        f"{BACKEND_URL}/reports/generate-instant",
        headers=headers,
        json=report_params_nov,
        timeout=60
    )

    if report_response_nov.status_code == 200:
        print(f"[OK] November report generated! Size: {len(report_response_nov.content)} bytes")

        filename_nov = f"test_november_{datetime.now().strftime('%H%M%S')}.xlsx"
        with open(filename_nov, 'wb') as f:
            f.write(report_response_nov.content)
        print(f"[OK] Saved as: {filename_nov}")

        # Try to read the Excel file
        try:
            import openpyxl
            wb = openpyxl.load_workbook(filename_nov)
            sheet = wb[wb.sheetnames[0]]

            print("\n[CHECKING NOVEMBER REPORT CONTENTS]")
            for row_idx in range(1, min(25, sheet.max_row + 1)):
                row = []
                for col_idx in range(1, min(3, sheet.max_column + 1)):
                    cell = sheet.cell(row_idx, col_idx)
                    if cell.value:
                        row.append(str(cell.value))
                if row:
                    print(f"  Row {row_idx}: {' | '.join(row)}")

            wb.close()
        except ImportError:
            print("[INFO] Install openpyxl to read Excel contents: pip install openpyxl")
        except Exception as e:
            print(f"[WARNING] Could not read Excel: {e}")

    else:
        print(f"[ERROR] November report generation failed: {report_response_nov.status_code}")
        print(f"Response: {report_response_nov.text}")
except Exception as e:
    print(f"[ERROR] November report request failed: {e}")

print("\n" + "=" * 80)
print("TEST COMPLETE")
print("=" * 80)
print("\nIf reports show $0.00:")
print("1. Check that MongoDB Atlas has data in salesHistory collection")
print("2. Verify backend is connecting to correct MongoDB")
print("3. Check that report_service.py fixes are deployed")
print("4. Look at backend logs on Render for errors")
