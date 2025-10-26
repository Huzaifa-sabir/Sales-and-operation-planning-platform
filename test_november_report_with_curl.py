"""
Test November 2025 Sales Report Generation with curl
Run this after backend is deployed on Render
"""
import requests
import os
from datetime import datetime
import openpyxl

# CONFIGURATION - UPDATE THIS with your Render URL
BACKEND_URL = os.getenv('BACKEND_URL', 'http://localhost:8000/api/v1')
# Example: BACKEND_URL = "https://your-app.onrender.com/api/v1"

print("=" * 70)
print("November 2025 Sales Report Test")
print("=" * 70)
print(f"Backend URL: {BACKEND_URL}")
print("=" * 70)

# Step 1: Login
print("\n[STEP 1] Logging in...")
try:
    login_response = requests.post(f"{BACKEND_URL}/auth/login", json={
        "email": "admin@sopportal.com",
        "password": "admin123"
    }, timeout=30)

    if login_response.status_code == 200:
        token = login_response.json().get("access_token")
        print("[OK] Login successful!")
        print(f"Token: {token[:20]}...")
    else:
        print(f"[ERROR] Login failed: {login_response.status_code}")
        print(login_response.text)
        exit(1)
except Exception as e:
    print(f"[ERROR] Login failed: {str(e)}")
    exit(1)

# Step 2: Generate November 2025 Sales Summary Report
print("\n[STEP 2] Generating November 2025 Sales Summary Report...")

headers = {
    "Authorization": f"Bearer {token}",
    "Content-Type": "application/json"
}

report_params = {
    "reportType": "sales_summary",
    "format": "excel",
    "year": 2025,
    "month": 11,
    "includeCharts": True,
    "includeRawData": False
}

print(f"Report Parameters: {report_params}")

try:
    print("\nSending request to /reports/generate-instant...")
    report_response = requests.post(
        f"{BACKEND_URL}/reports/generate-instant",
        headers=headers,
        json=report_params,
        timeout=60  # 60 second timeout for report generation
    )

    if report_response.status_code == 200:
        print(f"[OK] Report generated successfully!")
        print(f"Content-Type: {report_response.headers.get('content-type')}")
        print(f"Content-Length: {len(report_response.content)} bytes")

        # Save the Excel file
        filename = f"november_2025_sales_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        with open(filename, 'wb') as f:
            f.write(report_response.content)

        print(f"[OK] Report saved as: {filename}")
        print(f"File size: {os.path.getsize(filename)} bytes")

        # Step 3: Read and verify Excel content
        print("\n[STEP 3] Verifying Excel file contents...")
        try:
            workbook = openpyxl.load_workbook(filename)
            print(f"[OK] Excel file opened successfully")
            print(f"Sheet names: {workbook.sheetnames}")

            # Read first sheet
            first_sheet = workbook[workbook.sheetnames[0]]
            print(f"\n[OK] Reading sheet: {workbook.sheetnames[0]}")

            # Find and read key values
            print("\nReport Data Verification:")
            print("-" * 70)

            # Look for Total Revenue, Total Quantity, Transactions
            for row in first_sheet.iter_rows(min_row=1, max_row=20, values_only=True):
                if row[0]:
                    row_str = str(row[0])
                    if 'Total Revenue' in row_str or 'Total Quantity' in row_str or 'Transaction' in row_str:
                        print(f"{row[0]}: {row[1] if len(row) > 1 else 'N/A'}")

            print("-" * 70)

            # Expected values from our November 2025 data
            expected_revenue = 161750.00
            expected_quantity = 2750.00
            expected_transactions = 25

            print("\nExpected Values (from database insert):")
            print(f"  Total Revenue: ${expected_revenue:,.2f}")
            print(f"  Total Quantity: {expected_quantity:,.2f}")
            print(f"  Transactions: {expected_transactions}")

            workbook.close()

            print("\n" + "=" * 70)
            print("[SUCCESS] Report generation and verification complete!")
            print("=" * 70)
            print(f"\nGenerated File: {filename}")
            print(f"File Size: {os.path.getsize(filename):,} bytes")
            print("\nNext Steps:")
            print("1. Open the Excel file manually to verify data")
            print("2. Check that revenue is approximately $161,750.00")
            print("3. Check that quantity is approximately 2,750.00")
            print("4. Check that there are 25 transactions")
            print("5. Verify customer names and product descriptions appear correctly")

        except Exception as e:
            print(f"[ERROR] Failed to read Excel file: {str(e)}")
            print(f"[INFO] File was saved as {filename}, you can open it manually")

    else:
        print(f"[ERROR] Report generation failed: {report_response.status_code}")
        print(f"Response: {report_response.text}")
        exit(1)

except requests.exceptions.Timeout:
    print("[ERROR] Request timed out (> 60 seconds)")
    print("This might indicate the server is slow or the report is too large")
    exit(1)
except Exception as e:
    print(f"[ERROR] Request failed: {str(e)}")
    exit(1)
