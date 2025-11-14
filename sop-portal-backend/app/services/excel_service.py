"""
Excel Import/Export Service
Handles Excel file generation, template creation, data import/export
"""
from typing import List, Dict, Any, Optional, BinaryIO
from datetime import datetime
from io import BytesIO
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.utils.excel_helpers import (
    format_header_row,
    auto_adjust_column_width,
    format_currency_column,
    format_date_column,
    safe_str,
    safe_float,
    safe_bool
)
from app.models.customer import CustomerLocation
from app.models.product import ProductGroup, ProductManufacturing, ProductPricing


class ExcelService:
    """Service for Excel import/export operations"""

    # ==================== TEMPLATE GENERATORS ====================

    @staticmethod
    def generate_customer_template() -> BytesIO:
        """
        Generate Excel template for customer import
        Matches the format expected by the import functionality
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Customers"

        # Header row (row 1) - Bold and formatted
        headers = [
            "Customer ID*",
            "Customer Name*",
            "S&OP Customer Name",
            "Sales Rep Name",
            "Contact Person",
            "Contact Email",
            "Contact Phone",
            "Address Line 1",
            "City",
            "State",
            "Country",
            "Postal Code",
            "Payment Terms",
            "Credit Limit",
            "Is Active*"
        ]
        ws.append(headers)

        # Instructions row (row 2)
        instructions = [
            "Required - Unique ID (e.g., PATITO-000001)",
            "Required - Full customer name",
            "Optional - S&OP name (defaults to Customer Name)",
            "Optional - Sales rep name",
            "Optional - Contact person name",
            "Optional - Email address",
            "Optional - Phone number",
            "Optional - Street address",
            "Optional - City",
            "Optional - State/Province",
            "Optional - Country (default: USA)",
            "Optional - ZIP/Postal code",
            "Optional - Payment terms (e.g., Net 30)",
            "Optional - Credit limit (number)",
            "Required - true/false (default: true)"
        ]
        ws.append(instructions)

        # Sample data rows (rows 3-5)
        sample1 = [
            "PATITO-000001",
            "Industria Los Patitos, S.A.",
            "Industria Los Patitos",
            "JR",
            "John Doe",
            "john@patito.com",
            "+1-305-555-0100",
            "123 Main Street",
            "Miami",
            "FL",
            "USA",
            "33101",
            "Net 30",
            "50000.00",
            "true"
        ]
        ws.append(sample1)

        sample2 = [
            "CANADA-000002",
            "Canadawide Fruit Wholesalers Inc.",
            "Canadawide",
            "Anthony",
            "Jane Smith",
            "jane@canadawide.com",
            "+1-416-555-0200",
            "456 Business Blvd",
            "Toronto",
            "ON",
            "Canada",
            "M5H 2N2",
            "Net 45",
            "75000.00",
            "true"
        ]
        ws.append(sample2)

        sample3 = [
            "AAORG-000003",
            "A&A Organic Farms Corp.",
            "A&A Organic",
            "Randell",
            "Bob Johnson",
            "bob@aaorganic.com",
            "+1-310-555-0300",
            "789 Farm Road",
            "Los Angeles",
            "CA",
            "USA",
            "90001",
            "Net 30",
            "30000.00",
            "true"
        ]
        ws.append(sample3)

        # Format header row
        format_header_row(ws, 1)
        
        # Format instructions row (gray background)
        from openpyxl.styles import PatternFill, Font
        gray_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        for cell in ws[2]:
            cell.fill = gray_fill
            cell.font = Font(size=9, italic=True)

        # Format sample rows
        for row_idx in range(3, 6):
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if col_idx == 14:  # Credit Limit column
                    try:
                        cell.value = float(cell.value)
                    except:
                        pass

        # Format currency column (Credit Limit)
        format_currency_column(ws, "N", 3)
        
        # Auto-adjust column widths
        auto_adjust_column_width(ws)

        # Add note at bottom
        note_row = ws.max_row + 2
        ws.cell(row=note_row, column=1).value = "Note:"
        ws.cell(row=note_row, column=2).value = "Rows 3-5 are sample data. Delete these rows and add your customer data."
        ws.cell(row=note_row, column=2).font = Font(italic=True, color="666666")
        ws.merge_cells(f'B{note_row}:O{note_row}')

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def generate_product_template() -> BytesIO:
        """Generate Excel template for product import"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Products"

        # Headers
        headers = [
            "Item Code*",
            "Description*",
            "Group Code*",
            "Group Subgroup",
            "Group Description",
            "Manufacturing Location*",
            "Manufacturing Line",
            "Weight",
            "Unit of Measure*",
            "Average Price*",
            "Cost Price",
            "Currency"
        ]
        ws.append(headers)

        # Add sample row
        sample = [
            "110001",
            "Peeled Garlic 12x1 LB",
            "G1",
            "G1S7",
            "Group 1-2",
            "Miami",
            "Peeled Garlic Repack",
            "12.0",
            "CS",
            "52.00",
            "45.00",
            "USD"
        ]
        ws.append(sample)

        # Add instructions row
        ws.append([
            "Required (unique)",
            "Required",
            "Required",
            "Optional",
            "Optional",
            "Required",
            "Optional",
            "Optional (number)",
            "Required",
            "Required (number)",
            "Optional (number)",
            "Optional (default: USD)"
        ])

        # Format
        format_header_row(ws, 1)
        format_currency_column(ws, "J", 2)
        format_currency_column(ws, "K", 2)
        auto_adjust_column_width(ws)

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def generate_matrix_template() -> BytesIO:
        """Generate Excel template for product-customer matrix import"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Product-Customer Matrix"

        # Headers
        headers = [
            "Customer ID*",
            "Product Item Code*",
            "Customer Price",
            "Minimum Order Qty",
            "Maximum Order Qty",
            "Lead Time (Days)"
        ]
        ws.append(headers)

        # Add sample rows
        samples = [
            ["PATITO-000001", "110001", "50.00", "1", "100", "7"],
            ["PATITO-000001", "110002", "48.00", "1", "50", "7"],
            ["FOOD-000002", "110001", "52.00", "2", "200", "5"]
        ]
        for sample in samples:
            ws.append(sample)

        # Add instructions row
        ws.append([
            "Required",
            "Required",
            "Optional (number)",
            "Optional (number)",
            "Optional (number)",
            "Optional (number)"
        ])

        # Format
        format_header_row(ws, 1)
        format_currency_column(ws, "C", 2)
        auto_adjust_column_width(ws)

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    # ==================== EXPORT FUNCTIONS ====================

    @staticmethod
    def export_customers(customers: List[Dict[str, Any]]) -> BytesIO:
        """Export customers to Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Customers"

        # Headers
        headers = [
            "Customer ID",
            "Customer Name",
            "Contact Person",
            "Contact Email",
            "Contact Phone",
            "Address",
            "City",
            "State",
            "Country",
            "Postal Code",
            "Payment Terms",
            "Credit Limit",
            "Active",
            "Created At"
        ]
        ws.append(headers)

        # Data rows
        for customer in customers:
            location = customer.get("location", {}) or {}
            row = [
                customer.get("customerId", ""),
                customer.get("customerName", ""),
                customer.get("contactPerson", ""),
                customer.get("contactEmail", ""),
                customer.get("contactPhone", ""),
                location.get("address", ""),
                location.get("city", ""),
                location.get("state", ""),
                location.get("country", ""),
                location.get("postalCode", ""),
                customer.get("paymentTerms", ""),
                customer.get("creditLimit"),
                "Yes" if customer.get("isActive", True) else "No",
                customer.get("createdAt")
            ]
            ws.append(row)

        # Format
        format_header_row(ws, 1)
        format_currency_column(ws, "L", 2)
        format_date_column(ws, "N", 2)
        auto_adjust_column_width(ws)

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def export_products(products: List[Dict[str, Any]]) -> BytesIO:
        """Export products to Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Products"

        # Headers
        headers = [
            "Item Code",
            "Description",
            "Group Code",
            "Group Subgroup",
            "Group Description",
            "Manufacturing Location",
            "Manufacturing Line",
            "Weight",
            "Unit of Measure",
            "Average Price",
            "Cost Price",
            "Currency",
            "Active",
            "Created At"
        ]
        ws.append(headers)

        # Data rows
        for product in products:
            group = product.get("group", {}) or {}
            manufacturing = product.get("manufacturing", {}) or {}
            pricing = product.get("pricing", {}) or {}

            row = [
                product.get("itemCode", ""),
                product.get("description", ""),
                group.get("code", ""),
                group.get("subgroup", ""),
                group.get("desc", ""),
                manufacturing.get("location", ""),
                manufacturing.get("line", ""),
                product.get("weight"),
                product.get("uom", ""),
                pricing.get("avgPrice"),
                pricing.get("costPrice"),
                pricing.get("currency", "USD"),
                "Yes" if product.get("isActive", True) else "No",
                product.get("createdAt")
            ]
            ws.append(row)

        # Format
        format_header_row(ws, 1)
        format_currency_column(ws, "J", 2)
        format_currency_column(ws, "K", 2)
        format_date_column(ws, "N", 2)
        auto_adjust_column_width(ws)

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def export_sales_history(records: List[Dict[str, Any]]) -> BytesIO:
        """Export sales history records to Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Sales History"

        # Headers aligned with frontend expectations
        headers = [
            "Customer Name",
            "Customer ID",
            "Sales Rep",
            "Product Code",
            "Product Description",
            "Year-Month",
            "Year",
            "Month",
            "Quantity",
            "Unit Price",
            "Total Sales",
            "COGS",
            "Gross Profit",
            "GP %",
            "Created At",
        ]
        ws.append(headers)

        for rec in records:
            row = [
                rec.get("customerName", ""),
                rec.get("customerId", ""),
                rec.get("salesRepName", ""),
                rec.get("productCode", ""),
                rec.get("productDescription", ""),
                rec.get("yearMonth") or rec.get("month"),
                rec.get("year"),
                rec.get("month"),
                rec.get("quantity"),
                rec.get("unitPrice"),
                rec.get("totalSales"),
                rec.get("cogs") if rec.get("cogs") is not None else rec.get("costPrice"),
                rec.get("grossProfit"),
                rec.get("grossProfitPercent"),
                rec.get("createdAt"),
            ]
            ws.append(row)

        # Basic formatting
        format_header_row(ws, 1)
        format_currency_column(ws, "J", 2)  # Unit Price
        format_currency_column(ws, "K", 2)  # Total Sales
        format_currency_column(ws, "L", 2)  # COGS
        format_currency_column(ws, "M", 2)  # Gross Profit
        auto_adjust_column_width(ws)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    # ==================== IMPORT FUNCTIONS ====================

    @staticmethod
    def import_customers(file: BinaryIO) -> Dict[str, Any]:
        """
        Import customers from Excel file
        Returns dict with success/error information
        """
        try:
            df = pd.read_excel(file, sheet_name=0)
        except Exception as e:
            return {
                "success": False,
                "error": f"Failed to read Excel file: {str(e)}",
                "imported": [],
                "errors": []
            }

        imported = []
        errors = []

        for index, row in df.iterrows():
            row_num = index + 2  # +2 because of header and 0-index

            try:
                # Skip instruction rows and empty rows
                if pd.isna(row.iloc[0]) or str(row.iloc[0]).startswith("Required"):
                    continue

                # Required fields - try header names first, then fallback to column index
                customer_id = safe_str(row.get("Customer ID*") or row.iloc[0] if len(row) > 0 else None)
                customer_name = safe_str(row.get("Customer Name*") or row.iloc[1] if len(row) > 1 else None)

                if not customer_id or not customer_name:
                    errors.append({
                        "row": row_num,
                        "error": "Customer ID and Customer Name are required"
                    })
                    continue

                # Optional fields
                sop_customer_name = safe_str(row.get("S&OP Customer Name") or (row.iloc[2] if len(row) > 2 else None))
                sales_rep_name = safe_str(row.get("Sales Rep Name") or (row.iloc[3] if len(row) > 3 else None))
                
                # Build location object if any location data exists
                location = None
                # Try header names first, then column indices (accounting for new columns)
                address = safe_str(row.get("Address Line 1") or row.get("Address") or (row.iloc[7] if len(row) > 7 else None))
                city = safe_str(row.get("City") or (row.iloc[8] if len(row) > 8 else None))
                state = safe_str(row.get("State") or (row.iloc[9] if len(row) > 9 else None))
                country = safe_str(row.get("Country") or (row.iloc[10] if len(row) > 10 else None))
                postal_code = safe_str(row.get("Postal Code") or (row.iloc[11] if len(row) > 11 else None))

                if any([address, city, state, country, postal_code]):
                    location = {
                        "address": address,  # Note: CustomerLocation model uses "address", not "address1"
                        "city": city,
                        "state": state,
                        "country": country,
                        "zipCode": postal_code  # Note: Model uses zipCode, not postalCode
                    }

                # Is Active - default to True if not specified
                is_active_str = safe_str(row.get("Is Active*") or row.get("Is Active") or (row.iloc[14] if len(row) > 14 else None))
                is_active = True  # Default
                if is_active_str:
                    is_active_str_lower = str(is_active_str).lower().strip()
                    is_active = is_active_str_lower in ['true', 'yes', '1', 'active', 'y']

                customer_data = {
                    "customerId": customer_id,
                    "customerName": customer_name,
                    "sopCustomerName": sop_customer_name or customer_name,  # Default to customer_name if not provided
                    "salesRepName": sales_rep_name or "Default Sales Rep",  # Default if not provided
                    "location": location,
                    "contactPerson": safe_str(row.get("Contact Person") or (row.iloc[4] if len(row) > 4 else None)),
                    "contactEmail": safe_str(row.get("Contact Email") or (row.iloc[5] if len(row) > 5 else None)),
                    "contactPhone": safe_str(row.get("Contact Phone") or (row.iloc[6] if len(row) > 6 else None)),
                    "paymentTerms": safe_str(row.get("Payment Terms") or (row.iloc[12] if len(row) > 12 else None)),
                    "creditLimit": safe_float(row.get("Credit Limit") or (row.iloc[13] if len(row) > 13 else None)),
                    "isActive": is_active
                }

                imported.append(customer_data)

            except Exception as e:
                errors.append({
                    "row": row_num,
                    "error": str(e)
                })

        return {
            "success": len(errors) == 0,
            "imported": imported,
            "errors": errors,
            "totalRows": len(df),
            "successCount": len(imported),
            "errorCount": len(errors)
        }

    @staticmethod
    def import_products(file: BinaryIO) -> Dict[str, Any]:
        """
        Import products from Excel file
        Returns dict with success/error information
        """
        try:
            df = pd.read_excel(file, sheet_name=0)
        except Exception as e:
            return {
                "success": False,
                "error": f"Failed to read Excel file: {str(e)}",
                "imported": [],
                "errors": []
            }

        imported = []
        errors = []

        for index, row in df.iterrows():
            row_num = index + 2

            try:
                # Skip instruction rows and empty rows
                if pd.isna(row.iloc[0]) or str(row.iloc[0]).startswith("Required"):
                    continue

                # Required fields
                item_code = safe_str(row.get("Item Code*") or row.iloc[0])
                description = safe_str(row.get("Description*") or row.iloc[1])
                group_code = safe_str(row.get("Group Code*") or row.iloc[2])
                mfg_location = safe_str(row.get("Manufacturing Location*") or row.iloc[5])
                uom = safe_str(row.get("Unit of Measure*") or row.iloc[8])
                avg_price = safe_float(row.get("Average Price*") or row.iloc[9])

                if not all([item_code, description, group_code, mfg_location, uom]):
                    errors.append({
                        "row": row_num,
                        "error": "Item Code, Description, Group Code, Manufacturing Location, and UOM are required"
                    })
                    continue

                if avg_price is None:
                    errors.append({
                        "row": row_num,
                        "error": "Average Price is required"
                    })
                    continue

                # Build nested objects
                group = {
                    "code": group_code,
                    "subgroup": safe_str(row.get("Group Subgroup", row.iloc[3] if len(row) > 3 else None)),
                    "desc": safe_str(row.get("Group Description", row.iloc[4] if len(row) > 4 else None))
                }

                manufacturing = {
                    "location": mfg_location,
                    "line": safe_str(row.get("Manufacturing Line", row.iloc[6] if len(row) > 6 else None))
                }

                pricing = {
                    "avgPrice": avg_price,
                    "costPrice": safe_float(row.get("Cost Price", row.iloc[10] if len(row) > 10 else None)),
                    "currency": safe_str(row.get("Currency", row.iloc[11] if len(row) > 11 else "USD")) or "USD"
                }

                product_data = {
                    "itemCode": item_code,
                    "description": description,
                    "group": group,
                    "manufacturing": manufacturing,
                    "pricing": pricing,
                    "weight": safe_float(row.get("Weight", row.iloc[7] if len(row) > 7 else None)),
                    "uom": uom
                }

                imported.append(product_data)

            except Exception as e:
                errors.append({
                    "row": row_num,
                    "error": str(e)
                })

        return {
            "success": len(errors) == 0,
            "imported": imported,
            "errors": errors,
            "totalRows": len(df),
            "successCount": len(imported),
            "errorCount": len(errors)
        }

    @staticmethod
    def import_matrix(file: BinaryIO) -> Dict[str, Any]:
        """
        Import product-customer matrix from Excel file
        Returns dict with success/error information
        """
        try:
            df = pd.read_excel(file, sheet_name=0)
        except Exception as e:
            return {
                "success": False,
                "error": f"Failed to read Excel file: {str(e)}",
                "imported": [],
                "errors": []
            }

        imported = []
        errors = []

        for index, row in df.iterrows():
            row_num = index + 2

            try:
                # Skip instruction rows and empty rows
                if pd.isna(row.iloc[0]) or str(row.iloc[0]).startswith("Required"):
                    continue

                # Required fields
                customer_id = safe_str(row.get("Customer ID*") or row.iloc[0])
                product_id = safe_str(row.get("Product Item Code*") or row.iloc[1])

                if not customer_id or not product_id:
                    errors.append({
                        "row": row_num,
                        "error": "Customer ID and Product Item Code are required"
                    })
                    continue

                matrix_data = {
                    "customerId": customer_id,
                    "productId": product_id,
                    "customerPrice": safe_float(row.get("Customer Price", row.iloc[2] if len(row) > 2 else None)),
                    "minimumOrderQty": safe_float(row.get("Minimum Order Qty", row.iloc[3] if len(row) > 3 else None)),
                    "maximumOrderQty": safe_float(row.get("Maximum Order Qty", row.iloc[4] if len(row) > 4 else None)),
                    "leadTimeDays": safe_float(row.get("Lead Time (Days)", row.iloc[5] if len(row) > 5 else None))
                }

                imported.append(matrix_data)

            except Exception as e:
                errors.append({
                    "row": row_num,
                    "error": str(e)
                })

        return {
            "success": len(errors) == 0,
            "imported": imported,
            "errors": errors,
            "totalRows": len(df),
            "successCount": len(imported),
            "errorCount": len(errors)
        }
