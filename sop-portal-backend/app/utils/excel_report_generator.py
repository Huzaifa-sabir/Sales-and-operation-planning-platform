"""
Excel Report Generator with Charts
Generates professional Excel reports with formatting and visualizations
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference, Series
from openpyxl.utils import get_column_letter
from typing import Dict, Any, List
from datetime import datetime
import os


class ExcelReportGenerator:
    """Generate Excel reports with professional formatting and charts"""

    def __init__(self):
        self.workbook = None
        self.current_sheet = None

        # Define color scheme
        self.colors = {
            "header": "366092",  # Dark blue
            "subheader": "5B9BD5",  # Light blue
            "highlight": "FFC000",  # Orange
            "positive": "70AD47",  # Green
            "negative": "FF0000"   # Red
        }

    def create_workbook(self) -> openpyxl.Workbook:
        """Create new workbook"""
        self.workbook = openpyxl.Workbook()
        # Remove default sheet
        if "Sheet" in self.workbook.sheetnames:
            del self.workbook["Sheet"]
        return self.workbook

    def add_sheet(self, sheet_name: str):
        """Add new sheet"""
        self.current_sheet = self.workbook.create_sheet(sheet_name)
        return self.current_sheet

    def set_column_widths(self, widths: Dict[str, int]):
        """Set column widths"""
        for col_letter, width in widths.items():
            self.current_sheet.column_dimensions[col_letter].width = width

    def write_title(self, title: str, row: int = 1):
        """Write report title"""
        cell = self.current_sheet.cell(row=row, column=1, value=title)
        cell.font = Font(size=16, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=self.colors["header"], end_color=self.colors["header"], fill_type="solid")
        cell.alignment = Alignment(horizontal="left", vertical="center")

        # Merge cells for title
        self.current_sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)

    def write_metadata(self, metadata: Dict[str, Any], start_row: int = 3):
        """Write report metadata"""
        row = start_row
        for key, value in metadata.items():
            self.current_sheet.cell(row=row, column=1, value=key).font = Font(bold=True)
            self.current_sheet.cell(row=row, column=2, value=str(value))
            row += 1
        return row

    def write_table_header(self, headers: List[str], start_row: int, start_col: int = 1):
        """Write table header with formatting"""
        for col_idx, header in enumerate(headers, start=start_col):
            cell = self.current_sheet.cell(row=start_row, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=self.colors["subheader"], end_color=self.colors["subheader"], fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = self._get_border()

    def write_table_data(self, data: List[List[Any]], start_row: int, start_col: int = 1):
        """Write table data"""
        for row_idx, row_data in enumerate(data, start=start_row):
            for col_idx, value in enumerate(row_data, start=start_col):
                cell = self.current_sheet.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self._get_border(thin=True)
                cell.alignment = Alignment(horizontal="left", vertical="center")

                # Format numbers
                if isinstance(value, (int, float)):
                    cell.number_format = '#,##0.00'

    def add_bar_chart(
        self,
        title: str,
        categories_range: str,
        data_range: str,
        position: str = "H2",
        width: int = 15,
        height: int = 10
    ):
        """Add bar chart to current sheet"""
        try:
            chart = BarChart()
            chart.title = title
            chart.style = 10
            chart.width = width
            chart.height = height

            # Set data
            data = Reference(self.current_sheet, range_string=data_range)
            cats = Reference(self.current_sheet, range_string=categories_range)

            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)

            self.current_sheet.add_chart(chart, position)
        except Exception as e:
            # Skip chart generation if there's an error
            print(f"Warning: Could not generate chart '{title}': {e}")
            pass

    def add_line_chart(
        self,
        title: str,
        categories_range: str,
        data_ranges: List[str],
        position: str = "H2",
        width: int = 15,
        height: int = 10
    ):
        """Add line chart to current sheet"""
        try:
            chart = LineChart()
            chart.title = title
            chart.style = 10
            chart.width = width
            chart.height = height

            # Add multiple data series
            for data_range in data_ranges:
                data = Reference(self.current_sheet, range_string=data_range)
                chart.add_data(data, titles_from_data=True)

            cats = Reference(self.current_sheet, range_string=categories_range)
            chart.set_categories(cats)

            self.current_sheet.add_chart(chart, position)
        except Exception as e:
            # Skip chart generation if there's an error
            print(f"Warning: Could not generate chart '{title}': {e}")
            pass

    def auto_fit_columns(self):
        """Auto-fit column widths based on content"""
        for column in self.current_sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)
            self.current_sheet.column_dimensions[column_letter].width = adjusted_width

    def save(self, file_path: str):
        """Save workbook to file"""
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        self.workbook.save(file_path)

    def _get_border(self, thin: bool = False):
        """Get border style"""
        style = "thin" if thin else "medium"
        side = Side(style=style)
        return Border(left=side, right=side, top=side, bottom=side)

    # ==========================================
    # SPECIALIZED REPORT GENERATORS
    # ==========================================

    def generate_sales_summary_excel(self, data: Dict[str, Any], file_path: str):
        """Generate Sales Summary Excel Report"""
        self.create_workbook()

        # ===== SUMMARY SHEET =====
        summary_sheet = self.add_sheet("Summary")

        self.write_title("Sales Summary Report", row=1)

        # Metadata
        metadata = {
            "Generated At": data["generatedAt"],
            "Report Type": data["reportType"]
        }
        self.write_metadata(metadata, start_row=3)

        # Overall summary
        row = 7
        summary_sheet.cell(row=row, column=1, value="Overall Statistics").font = Font(size=14, bold=True)
        row += 1

        summary_data = data["summary"]
        stats = [
            ["Total Revenue", f"${summary_data['totalRevenue']:,.2f}"],
            ["Total Quantity", f"{summary_data['totalQuantity']:,.2f}"],
            ["Transactions", f"{summary_data['transactionCount']:,}"],
            ["Avg Quantity/Transaction", f"{summary_data['avgQuantity']:,.2f}"],
            ["Avg Unit Price", f"${summary_data['avgUnitPrice']:,.2f}"]
        ]

        for stat in stats:
            summary_sheet.cell(row=row, column=1, value=stat[0]).font = Font(bold=True)
            summary_sheet.cell(row=row, column=2, value=stat[1])
            row += 1

        self.set_column_widths({"A": 30, "B": 20})

        # ===== MONTHLY TRENDS SHEET =====
        trends_sheet = self.add_sheet("Monthly Trends")
        self.current_sheet = trends_sheet

        self.write_title("Monthly Sales Trends", row=1)

        headers = ["Month", "Quantity", "Revenue", "Transactions"]
        self.write_table_header(headers, start_row=3)

        trends_data = [
            [
                trend["monthLabel"],
                trend["quantity"],
                trend["revenue"],
                trend["transactions"]
            ]
            for trend in data["monthlyTrends"]
        ]

        self.write_table_data(trends_data, start_row=4)

        # Add line chart
        if len(trends_data) > 0:
            chart_row = len(trends_data) + 4
            self.add_line_chart(
                title="Sales Trend",
                categories_range=f"A4:A{chart_row}",
                data_ranges=[f"B4:B{chart_row}", f"C4:C{chart_row}"],
                position=f"E3"
            )

        self.auto_fit_columns()

        # ===== TOP CUSTOMERS SHEET =====
        customers_sheet = self.add_sheet("Top Customers")
        self.current_sheet = customers_sheet

        self.write_title("Top 10 Customers by Revenue", row=1)

        headers = ["Customer ID", "Customer Name", "Revenue", "Quantity", "Transactions"]
        self.write_table_header(headers, start_row=3)

        customers_data = [
            [
                c["customerId"],
                c["customerName"],
                c["totalRevenue"],
                c["totalQuantity"],
                c["transactions"]
            ]
            for c in data["topCustomers"]
        ]

        self.write_table_data(customers_data, start_row=4)

        # Add bar chart
        if len(customers_data) > 0:
            self.add_bar_chart(
                title="Top Customers Revenue",
                categories_range=f"B4:B{3+len(customers_data)}",
                data_range=f"C3:C{3+len(customers_data)}",
                position="G3"
            )

        self.auto_fit_columns()

        # ===== TOP PRODUCTS SHEET =====
        products_sheet = self.add_sheet("Top Products")
        self.current_sheet = products_sheet

        self.write_title("Top 10 Products by Volume", row=1)

        headers = ["Product ID", "Description", "Quantity", "Revenue", "Transactions"]
        self.write_table_header(headers, start_row=3)

        products_data = [
            [
                p["productId"],
                p["productDescription"],
                p["totalQuantity"],
                p["totalRevenue"],
                p["transactions"]
            ]
            for p in data["topProducts"]
        ]

        self.write_table_data(products_data, start_row=4)

        # Add bar chart
        if len(products_data) > 0:
            self.add_bar_chart(
                title="Top Products Volume",
                categories_range=f"B4:B{3+len(products_data)}",
                data_range=f"C3:C{3+len(products_data)}",
                position="G3"
            )

        self.auto_fit_columns()

        # Save workbook
        self.save(file_path)

    def generate_forecast_vs_actual_excel(self, data: Dict[str, Any], file_path: str):
        """Generate Forecast vs Actual Excel Report"""
        self.create_workbook()

        # ===== SUMMARY SHEET =====
        summary_sheet = self.add_sheet("Summary")
        self.write_title("Forecast vs Actual Analysis", row=1)

        metadata = {
            "Generated At": data["generatedAt"],
            "Report Type": data["reportType"]
        }
        self.write_metadata(metadata, start_row=3)

        # Overall summary
        row = 7
        summary_sheet.cell(row=row, column=1, value="Overall Variance").font = Font(size=14, bold=True)
        row += 1

        summary_data = data["summary"]
        stats = [
            ["Total Forecast", f"{summary_data['totalForecast']:,.2f}"],
            ["Total Actual", f"{summary_data['totalActual']:,.2f}"],
            ["Overall Variance", f"{summary_data['overallVariance']:,.2f}"],
            ["Variance %", f"{summary_data['overallVariancePercent']:.2f}%"],
            ["Accuracy Rate (±10%)", f"{summary_data['accuracyRate']:.2f}%"],
            ["Records Analyzed", f"{summary_data['recordCount']:,}"]
        ]

        for stat in stats:
            summary_sheet.cell(row=row, column=1, value=stat[0]).font = Font(bold=True)
            cell = summary_sheet.cell(row=row, column=2, value=stat[1])

            # Color code variance
            if "Variance" in stat[0] and summary_data['overallVariance'] != 0:
                fill_color = self.colors["positive"] if summary_data['overallVariance'] > 0 else self.colors["negative"]
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

            row += 1

        self.set_column_widths({"A": 30, "B": 20})

        # ===== VARIANCE DETAILS SHEET =====
        details_sheet = self.add_sheet("Variance Details")
        self.current_sheet = details_sheet

        self.write_title("Monthly Variance Details", row=1)

        headers = ["Customer", "Product", "Month", "Forecast", "Actual", "Variance", "Variance %", "Status"]
        self.write_table_header(headers, start_row=3)

        variance_data = [
            [
                v["customerId"],
                v["productId"],
                v["monthLabel"],
                v["forecastQuantity"],
                v["actualQuantity"],
                v["variance"],
                v["variancePercent"],
                v["status"]
            ]
            for v in data["varianceDetails"]
        ]

        # Write data with conditional formatting
        for row_idx, row_data in enumerate(variance_data, start=4):
            for col_idx, value in enumerate(row_data, start=1):
                cell = details_sheet.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self._get_border(thin=True)

                # Color code status
                if col_idx == 8:  # Status column
                    if value == "Over":
                        cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                    elif value == "Under":
                        cell.fill = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")

        self.auto_fit_columns()

        # Save workbook
        self.save(file_path)

    def generate_monthly_dashboard_excel(self, data: Dict[str, Any], file_path: str):
        """Generate Monthly Dashboard Excel Report"""
        self.create_workbook()

        # ===== DASHBOARD SHEET =====
        dashboard_sheet = self.add_sheet("Dashboard")
        self.write_title(f"Monthly Dashboard - {data['targetPeriod']}", row=1)

        row = 3

        # Current Month KPIs
        dashboard_sheet.cell(row=row, column=1, value="Current Month Performance").font = Font(size=14, bold=True)
        row += 1

        current_month = data["currentMonth"]
        kpis = [
            ["Revenue", f"${current_month['revenue']:,.2f}"],
            ["Quantity Sold", f"{current_month['quantity']:,.2f}"],
            ["Transactions", f"{current_month['transactions']:,}"]
        ]

        for kpi in kpis:
            dashboard_sheet.cell(row=row, column=1, value=kpi[0]).font = Font(bold=True)
            cell = dashboard_sheet.cell(row=row, column=2, value=kpi[1])
            cell.font = Font(size=12, bold=True, color=self.colors["positive"])
            row += 1

        row += 2

        # YTD Performance
        dashboard_sheet.cell(row=row, column=1, value="Year-to-Date Performance").font = Font(size=14, bold=True)
        row += 1

        ytd = data["yearToDate"]
        ytd_kpis = [
            ["YTD Revenue", f"${ytd['revenue']:,.2f}"],
            ["YTD Quantity", f"{ytd['quantity']:,.2f}"]
        ]

        for kpi in ytd_kpis:
            dashboard_sheet.cell(row=row, column=1, value=kpi[0]).font = Font(bold=True)
            dashboard_sheet.cell(row=row, column=2, value=kpi[1])
            row += 1

        row += 2

        # Top Customers
        dashboard_sheet.cell(row=row, column=1, value="Top 5 Customers This Month").font = Font(size=14, bold=True)
        row += 1

        headers = ["Customer", "Revenue"]
        for col_idx, header in enumerate(headers, start=1):
            cell = dashboard_sheet.cell(row=row, column=col_idx, value=header)
            cell.font = Font(bold=True)
        row += 1

        for customer in data["topCustomers"]:
            dashboard_sheet.cell(row=row, column=1, value=customer["customerName"])
            dashboard_sheet.cell(row=row, column=2, value=f"${customer['revenue']:,.2f}")
            row += 1

        row += 2

        # Top Products
        dashboard_sheet.cell(row=row, column=1, value="Top 5 Products This Month").font = Font(size=14, bold=True)
        row += 1

        headers = ["Product", "Quantity"]
        for col_idx, header in enumerate(headers, start=1):
            cell = dashboard_sheet.cell(row=row, column=col_idx, value=header)
            cell.font = Font(bold=True)
        row += 1

        for product in data["topProducts"]:
            dashboard_sheet.cell(row=row, column=1, value=product["productDescription"])
            dashboard_sheet.cell(row=row, column=2, value=f"{product['quantity']:,.2f}")
            row += 1

        row += 2

        # Forecast Submission Status
        dashboard_sheet.cell(row=row, column=1, value="Forecast Submission Status").font = Font(size=14, bold=True)
        row += 1

        submission = data["forecastSubmission"]
        if "cycleName" in submission:
            sub_stats = [
                ["Active Cycle", submission.get("cycleName", "N/A")],
                ["Submission Rate", f"{submission.get('submissionRate', 0):.2f}%"],
                ["Submitted", f"{submission.get('submittedForecasts', 0)}/{submission.get('totalForecasts', 0)}"]
            ]

            for stat in sub_stats:
                dashboard_sheet.cell(row=row, column=1, value=stat[0]).font = Font(bold=True)
                dashboard_sheet.cell(row=row, column=2, value=stat[1])
                row += 1

        self.set_column_widths({"A": 30, "B": 25, "C": 15})

        # Save workbook
        self.save(file_path)

    def generate_customer_performance_excel(self, data: Dict[str, Any], file_path: str):
        """Generate Customer Performance Excel Report"""
        self.create_workbook()

        # ===== SUMMARY SHEET =====
        summary_sheet = self.add_sheet("Summary")
        self.write_title("Customer Performance Analysis", row=1)

        metadata = {
            "Generated At": data["generatedAt"],
            "Total Customers": data["summary"]["totalCustomers"],
            "Total Revenue": f"${data['summary']['totalRevenue']:,.2f}",
            "Avg Revenue/Customer": f"${data['summary']['avgRevenuePerCustomer']:,.2f}"
        }
        self.write_metadata(metadata, start_row=3)

        # ===== CUSTOMER DETAILS SHEET =====
        details_sheet = self.add_sheet("Customer Details")
        self.current_sheet = details_sheet

        self.write_title("Top 50 Customers", row=1)

        headers = ["Customer ID", "Name", "Region", "Territory", "Revenue", "Quantity",
                   "Transactions", "Avg Order Value", "Product Diversity", "Revenue %", "Last Purchase"]
        self.write_table_header(headers, start_row=3)

        customer_data = [
            [
                c["customerId"],
                c["customerName"],
                c["region"],
                c["territory"],
                c["totalRevenue"],
                c["totalQuantity"],
                c["transactions"],
                c["avgOrderValue"],
                c["productDiversity"],
                c["revenueContribution"],
                c["lastPurchase"]
            ]
            for c in data["customers"]
        ]

        self.write_table_data(customer_data, start_row=4)

        # Add bar chart for top 10
        if len(customer_data) >= 10:
            self.add_bar_chart(
                title="Top 10 Customers by Revenue",
                categories_range=f"B4:B13",
                data_range=f"E3:E13",
                position="M3"
            )

        self.auto_fit_columns()
        self.save(file_path)

    def generate_product_analysis_excel(self, data: Dict[str, Any], file_path: str):
        """Generate Product Analysis Excel Report"""
        self.create_workbook()

        # ===== SUMMARY SHEET =====
        summary_sheet = self.add_sheet("Summary")
        self.write_title("Product Analysis Report", row=1)

        metadata = {
            "Generated At": data["generatedAt"],
            "Total Products": data["summary"]["totalProducts"],
            "Total Revenue": f"${data['summary']['totalRevenue']:,.2f}",
            "Total Categories": data["summary"]["totalCategories"]
        }
        self.write_metadata(metadata, start_row=3)

        # ===== PRODUCT DETAILS SHEET =====
        details_sheet = self.add_sheet("Product Details")
        self.current_sheet = details_sheet

        self.write_title("Top 50 Products", row=1)

        headers = ["Product ID", "Description", "Category", "UOM",
                   "Quantity", "Revenue", "Transactions", "Avg Price", "Customer Reach"]
        self.write_table_header(headers, start_row=3)

        product_data = [
            [
                p["productId"],
                p["itemDescription"],
                p["category"],
                p["unitOfMeasure"],
                p["totalQuantity"],
                p["totalRevenue"],
                p["transactions"],
                p["avgPrice"],
                p["customerReach"]
            ]
            for p in data["products"]
        ]

        self.write_table_data(product_data, start_row=4)
        self.auto_fit_columns()

        # ===== CATEGORY BREAKDOWN SHEET =====
        category_sheet = self.add_sheet("Category Breakdown")
        self.current_sheet = category_sheet

        self.write_title("Sales by Category", row=1)

        headers = ["Category", "Revenue", "Quantity", "Products"]
        self.write_table_header(headers, start_row=3)

        category_data = [
            [
                c["category"],
                c["revenue"],
                c["quantity"],
                c["products"]
            ]
            for c in data["categoryBreakdown"]
        ]

        self.write_table_data(category_data, start_row=4)

        # Add pie chart for categories
        if len(category_data) > 0:
            self.add_bar_chart(
                title="Revenue by Category",
                categories_range=f"A4:A{3+len(category_data)}",
                data_range=f"B3:B{3+len(category_data)}",
                position="F3"
            )

        self.auto_fit_columns()
        self.save(file_path)

    def generate_cycle_submission_status_excel(self, data: Dict[str, Any], file_path: str):
        """Generate Cycle Submission Status Excel Report"""
        self.create_workbook()

        # ===== SUMMARY SHEET =====
        summary_sheet = self.add_sheet("Summary")
        self.write_title(f"Cycle Submission Status - {data['cycle']['cycleName']}", row=1)

        metadata = {
            "Generated At": data["generatedAt"],
            "Cycle ID": data["cycle"]["cycleId"],
            "Cycle Status": data["cycle"]["status"],
            "Start Date": data["cycle"]["startDate"],
            "End Date": data["cycle"]["endDate"]
        }
        self.write_metadata(metadata, start_row=3)

        row = 9
        summary_sheet.cell(row=row, column=1, value="Overall Statistics").font = Font(size=14, bold=True)
        row += 1

        summary_data = data["summary"]
        stats = [
            ["Total Forecasts", f"{summary_data['totalForecasts']:,}"],
            ["Submitted", f"{summary_data['submittedForecasts']:,}"],
            ["Draft", f"{summary_data['draftForecasts']:,}"],
            ["Submission Rate", f"{summary_data['submissionRate']:.2f}%"]
        ]

        for stat in stats:
            summary_sheet.cell(row=row, column=1, value=stat[0]).font = Font(bold=True)
            summary_sheet.cell(row=row, column=2, value=stat[1])
            row += 1

        self.set_column_widths({"A": 30, "B": 20})

        # ===== SALES REP BREAKDOWN SHEET =====
        breakdown_sheet = self.add_sheet("Sales Rep Breakdown")
        self.current_sheet = breakdown_sheet

        self.write_title("Sales Rep Submission Details", row=1)

        headers = ["Sales Rep ID", "Total Forecasts", "Submitted", "Draft",
                   "Submission Rate %", "Last Submission"]
        self.write_table_header(headers, start_row=3)

        rep_data = [
            [
                rep["salesRepId"],
                rep["totalForecasts"],
                rep["submitted"],
                rep["draft"],
                rep["submissionRate"],
                rep["lastSubmission"]
            ]
            for rep in data["salesRepBreakdown"]
        ]

        self.write_table_data(rep_data, start_row=4)
        self.auto_fit_columns()

        self.save(file_path)

    def generate_gross_profit_analysis_excel(self, data: Dict[str, Any], file_path: str):
        """Generate Gross Profit Analysis Excel Report"""
        self.create_workbook()

        # ===== SUMMARY SHEET =====
        summary_sheet = self.add_sheet("Summary")
        self.write_title("Gross Profit Analysis", row=1)

        metadata = {
            "Generated At": data["generatedAt"],
            "Report Type": data["reportType"]
        }
        self.write_metadata(metadata, start_row=3)

        row = 7
        summary_sheet.cell(row=row, column=1, value="Overall Profitability").font = Font(size=14, bold=True)
        row += 1

        summary_data = data["summary"]
        stats = [
            ["Total Revenue", f"${summary_data['totalRevenue']:,.2f}"],
            ["Total Cost", f"${summary_data['totalCost']:,.2f}"],
            ["Gross Profit", f"${summary_data['totalGrossProfit']:,.2f}"],
            ["Profit Margin", f"{summary_data['overallMargin']:.2f}%"]
        ]

        for stat in stats:
            summary_sheet.cell(row=row, column=1, value=stat[0]).font = Font(bold=True)
            cell = summary_sheet.cell(row=row, column=2, value=stat[1])

            # Highlight profit in green
            if "Profit" in stat[0]:
                cell.fill = PatternFill(start_color=self.colors["positive"],
                                       end_color=self.colors["positive"], fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
            row += 1

        self.set_column_widths({"A": 30, "B": 25})

        # ===== PROFIT DETAILS SHEET =====
        details_sheet = self.add_sheet("Profit Details")
        self.current_sheet = details_sheet

        self.write_title("Top 50 Profitable Items", row=1)

        headers = ["Customer ID", "Customer", "Product ID", "Product",
                   "Revenue", "Cost", "Gross Profit", "Margin %", "Quantity"]
        self.write_table_header(headers, start_row=3)

        profit_data = [
            [
                item["customerId"],
                item["customerName"],
                item["productId"],
                item["productDescription"],
                item["revenue"],
                item["cost"],
                item["grossProfit"],
                item["profitMargin"],
                item["quantity"]
            ]
            for item in data["profitDetails"]
        ]

        # Write data with conditional formatting
        for row_idx, row_data in enumerate(profit_data, start=4):
            for col_idx, value in enumerate(row_data, start=1):
                cell = details_sheet.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self._get_border(thin=True)

                # Format numbers
                if isinstance(value, (int, float)) and col_idx >= 5:
                    cell.number_format = '#,##0.00'

                # Color code margin column
                if col_idx == 8:  # Margin % column
                    if value >= 30:
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif value >= 20:
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    elif value < 10:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        self.auto_fit_columns()
        self.save(file_path)

    def generate_forecast_accuracy_excel(self, data: Dict[str, Any], file_path: str):
        """Generate Forecast Accuracy Excel Report"""
        self.create_workbook()

        # ===== SUMMARY SHEET =====
        summary_sheet = self.add_sheet("Summary")
        self.write_title("Forecast Accuracy Analysis", row=1)

        metadata = {
            "Generated At": data["generatedAt"],
            "Report Type": data["reportType"]
        }
        self.write_metadata(metadata, start_row=3)

        row = 7
        summary_sheet.cell(row=row, column=1, value="Overall Accuracy Metrics").font = Font(size=14, bold=True)
        row += 1

        summary_data = data["summary"]
        stats = [
            ["Total Comparisons", f"{summary_data['totalComparisons']:,}"],
            ["Overall MAPE", f"{summary_data['overallMAPE']:.2f}%"],
            ["Accuracy Rate (±10%)", f"{summary_data['accuracyRate']:.2f}%"],
            ["Accurate Forecasts", f"{summary_data['accurateForecasts']:,}"]
        ]

        for stat in stats:
            summary_sheet.cell(row=row, column=1, value=stat[0]).font = Font(bold=True)
            cell = summary_sheet.cell(row=row, column=2, value=stat[1])

            # Highlight accuracy rate
            if "Accuracy Rate" in stat[0]:
                accuracy = summary_data['accuracyRate']
                fill_color = self.colors["positive"] if accuracy >= 70 else self.colors["highlight"] if accuracy >= 50 else self.colors["negative"]
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
            row += 1

        self.set_column_widths({"A": 30, "B": 20})

        # ===== SALES REP PERFORMANCE SHEET =====
        performance_sheet = self.add_sheet("Sales Rep Performance")
        self.current_sheet = performance_sheet

        self.write_title("Forecast Accuracy by Sales Rep", row=1)

        headers = ["Sales Rep ID", "Total Forecasts", "Comparisons", "Avg MAPE %", "Accuracy Rate %"]
        self.write_table_header(headers, start_row=3)

        rep_data = [
            [
                rep["salesRepId"],
                rep["totalForecasts"],
                rep["comparisons"],
                rep["avgMAPE"],
                rep["accuracyRate"]
            ]
            for rep in data["salesRepPerformance"]
        ]

        self.write_table_data(rep_data, start_row=4)

        # Add bar chart
        if len(rep_data) > 0:
            self.add_bar_chart(
                title="Accuracy Rate by Sales Rep",
                categories_range=f"A4:A{3+min(len(rep_data), 10)}",
                data_range=f"E3:E{3+min(len(rep_data), 10)}",
                position="G3"
            )

        self.auto_fit_columns()

        # ===== ACCURACY DETAILS SHEET =====
        details_sheet = self.add_sheet("Accuracy Details")
        self.current_sheet = details_sheet

        self.write_title("Detailed Forecast Accuracy (Top 100)", row=1)

        headers = ["Sales Rep", "Customer", "Product", "Year", "Month",
                   "Forecast Qty", "Actual Qty", "APE %", "Accurate?"]
        self.write_table_header(headers, start_row=3)

        accuracy_data = [
            [
                detail["salesRepId"],
                detail["customerId"],
                detail["productId"],
                detail["year"],
                detail["month"],
                detail["forecastQty"],
                detail["actualQty"],
                detail["ape"],
                "Yes" if detail["isAccurate"] else "No"
            ]
            for detail in data["accuracyDetails"]
        ]

        # Write data with conditional formatting
        for row_idx, row_data in enumerate(accuracy_data, start=4):
            for col_idx, value in enumerate(row_data, start=1):
                cell = details_sheet.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self._get_border(thin=True)

                # Color code accurate column
                if col_idx == 9:  # Accurate column
                    if value == "Yes":
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        self.auto_fit_columns()
        self.save(file_path)
