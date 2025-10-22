"""
Excel Helper Utilities
Functions for formatting Excel files, handling dates, currency, etc.
"""
from datetime import datetime
from typing import Any, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def format_header_row(worksheet: Any, row: int = 1) -> None:
    """
    Format the header row with bold font, background color, and borders
    """
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for cell in worksheet[row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = alignment
        cell.border = thin_border


def auto_adjust_column_width(worksheet: Any, min_width: int = 10, max_width: int = 50) -> None:
    """
    Auto-adjust column widths based on content
    """
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass

        adjusted_width = min(max(max_length + 2, min_width), max_width)
        worksheet.column_dimensions[column_letter].width = adjusted_width


def format_currency_column(worksheet: Any, column_letter: str, start_row: int = 2) -> None:
    """
    Format a column as currency (USD)
    """
    for row in range(start_row, worksheet.max_row + 1):
        cell = worksheet[f"{column_letter}{row}"]
        if cell.value is not None:
            try:
                cell.number_format = '"$"#,##0.00'
            except:
                pass


def format_date_column(worksheet: Any, column_letter: str, start_row: int = 2) -> None:
    """
    Format a column as date (YYYY-MM-DD)
    """
    for row in range(start_row, worksheet.max_row + 1):
        cell = worksheet[f"{column_letter}{row}"]
        if cell.value is not None:
            try:
                cell.number_format = 'YYYY-MM-DD'
            except:
                pass


def format_datetime_column(worksheet: Any, column_letter: str, start_row: int = 2) -> None:
    """
    Format a column as datetime (YYYY-MM-DD HH:MM:SS)
    """
    for row in range(start_row, worksheet.max_row + 1):
        cell = worksheet[f"{column_letter}{row}"]
        if cell.value is not None:
            try:
                cell.number_format = 'YYYY-MM-DD HH:MM:SS'
            except:
                pass


def format_percentage_column(worksheet: Any, column_letter: str, start_row: int = 2) -> None:
    """
    Format a column as percentage
    """
    for row in range(start_row, worksheet.max_row + 1):
        cell = worksheet[f"{column_letter}{row}"]
        if cell.value is not None:
            try:
                cell.number_format = '0.00%'
            except:
                pass


def format_number_column(worksheet: Any, column_letter: str, start_row: int = 2, decimals: int = 0) -> None:
    """
    Format a column as number with specified decimal places
    """
    if decimals == 0:
        number_format = '#,##0'
    else:
        number_format = f'#,##0.{"0" * decimals}'

    for row in range(start_row, worksheet.max_row + 1):
        cell = worksheet[f"{column_letter}{row}"]
        if cell.value is not None:
            try:
                cell.number_format = number_format
            except:
                pass


def add_data_validation(worksheet: Any, cell_range: str, values: list) -> None:
    """
    Add dropdown data validation to a cell range
    """
    from openpyxl.worksheet.datavalidation import DataValidation

    dv = DataValidation(type="list", formula1=f'"{",".join(values)}"', allow_blank=True)
    dv.add(cell_range)
    worksheet.add_data_validation(dv)


def parse_excel_date(value: Any) -> Optional[datetime]:
    """
    Parse various date formats from Excel
    """
    if value is None or value == "":
        return None

    if isinstance(value, datetime):
        return value

    if isinstance(value, str):
        # Try common date formats
        date_formats = [
            "%Y-%m-%d",
            "%m/%d/%Y",
            "%d/%m/%Y",
            "%Y-%m-%d %H:%M:%S",
            "%m/%d/%Y %H:%M:%S"
        ]

        for fmt in date_formats:
            try:
                return datetime.strptime(value, fmt)
            except ValueError:
                continue

    return None


def safe_str(value: Any) -> str:
    """
    Safely convert value to string, handling None
    """
    if value is None or value == "":
        return ""
    return str(value).strip()


def safe_int(value: Any) -> Optional[int]:
    """
    Safely convert value to int, handling None and errors
    """
    if value is None or value == "":
        return None
    try:
        return int(float(value))
    except (ValueError, TypeError):
        return None


def safe_float(value: Any) -> Optional[float]:
    """
    Safely convert value to float, handling None and errors
    """
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def safe_bool(value: Any) -> bool:
    """
    Safely convert value to boolean
    """
    if value is None or value == "":
        return False

    if isinstance(value, bool):
        return value

    if isinstance(value, str):
        return value.lower() in ["true", "yes", "1", "y", "active"]

    return bool(value)
