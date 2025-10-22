"""
Forecast Excel Import Helper
Handles bulk forecast import from Excel files
"""
import openpyxl
from typing import List, Dict, Any, BinaryIO
from datetime import datetime
from fastapi import HTTPException, status

from app.models.forecast import MonthlyForecast, BulkForecastData


class ForecastExcelImporter:
    """Helper class for importing forecasts from Excel"""

    @staticmethod
    def parse_forecast_excel(file_content: BinaryIO, cycle_months: List[Dict[str, Any]]) -> List[BulkForecastData]:
        """
        Parse Excel file and extract forecast data

        Expected Excel format:
        Row 1: Headers (Customer ID, Product ID, Use Customer Price, Override Price, Notes, Month columns...)
        Row 2+: Data rows

        Month columns should be in YYYY-MM format matching the cycle's planning period
        """
        try:
            workbook = openpyxl.load_workbook(file_content, read_only=True)
            sheet = workbook.active

            # Read headers from first row
            headers = []
            for cell in sheet[1]:
                if cell.value:
                    headers.append(str(cell.value).strip())

            # Expected column names
            required_columns = ["Customer ID", "Product ID"]

            # Validate required columns exist
            for col in required_columns:
                if col not in headers:
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail=f"Missing required column: {col}"
                    )

            # Find column indices
            customer_id_idx = headers.index("Customer ID")
            product_id_idx = headers.index("Product ID")
            use_customer_price_idx = headers.index("Use Customer Price") if "Use Customer Price" in headers else None
            override_price_idx = headers.index("Override Price") if "Override Price" in headers else None
            notes_idx = headers.index("Notes") if "Notes" in headers else None

            # Find month columns (any column in YYYY-MM format)
            month_columns = {}
            for idx, header in enumerate(headers):
                # Try to parse as YYYY-MM
                try:
                    if "-" in header and len(header.split("-")) == 2:
                        year_str, month_str = header.split("-")
                        year = int(year_str)
                        month = int(month_str)
                        if 1 <= month <= 12:
                            month_label = f"{year}-{str(month).zfill(2)}"
                            month_columns[idx] = {
                                "year": year,
                                "month": month,
                                "monthLabel": month_label
                            }
                except:
                    continue

            if not month_columns:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="No valid month columns found. Month columns should be in YYYY-MM format."
                )

            # Parse data rows
            forecasts = []
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not row or not any(row):  # Skip empty rows
                    continue

                customer_id = str(row[customer_id_idx]).strip() if row[customer_id_idx] else None
                product_id = str(row[product_id_idx]).strip() if row[product_id_idx] else None

                if not customer_id or not product_id:
                    continue  # Skip rows without customer/product

                # Parse pricing options
                use_customer_price = True
                if use_customer_price_idx is not None and row[use_customer_price_idx] is not None:
                    val = str(row[use_customer_price_idx]).strip().lower()
                    use_customer_price = val in ["true", "yes", "1", "y"]

                override_price = None
                if override_price_idx is not None and row[override_price_idx] is not None:
                    try:
                        override_price = float(row[override_price_idx])
                    except:
                        pass

                notes = str(row[notes_idx]).strip() if notes_idx is not None and row[notes_idx] else None

                # Parse monthly quantities
                monthly_forecasts = []
                for col_idx, month_info in month_columns.items():
                    quantity = 0.0
                    if col_idx < len(row) and row[col_idx] is not None:
                        try:
                            quantity = float(row[col_idx])
                        except:
                            quantity = 0.0

                    # Find if this month is historical/current/future
                    is_historical = False
                    is_current = False
                    is_future = False

                    for cycle_month in cycle_months:
                        if (cycle_month["year"] == month_info["year"] and
                            cycle_month["month"] == month_info["month"]):
                            is_historical = cycle_month.get("isHistorical", False)
                            is_current = cycle_month.get("isCurrent", False)
                            is_future = cycle_month.get("isFuture", False)
                            break

                    monthly_forecast = MonthlyForecast(
                        year=month_info["year"],
                        month=month_info["month"],
                        monthLabel=month_info["monthLabel"],
                        quantity=quantity,
                        unitPrice=None,  # Will be set by service
                        revenue=None,  # Will be calculated by service
                        notes=None,
                        isHistorical=is_historical,
                        isCurrent=is_current,
                        isFuture=is_future
                    )
                    monthly_forecasts.append(monthly_forecast)

                # Create bulk forecast data
                bulk_data = BulkForecastData(
                    customerId=customer_id,
                    productId=product_id,
                    monthlyForecasts=monthly_forecasts,
                    useCustomerPrice=use_customer_price,
                    overridePrice=override_price,
                    notes=notes
                )
                forecasts.append(bulk_data)

            return forecasts

        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Error parsing Excel file: {str(e)}"
            )

    @staticmethod
    def generate_forecast_template(cycle_months: List[Dict[str, Any]]) -> openpyxl.Workbook:
        """
        Generate an Excel template for forecast import

        Returns a workbook with proper headers and formatting
        """
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Forecast Template"

        # Create headers
        headers = ["Customer ID", "Product ID", "Use Customer Price", "Override Price", "Notes"]

        # Add month columns (only future months for forecasting)
        future_months = [m for m in cycle_months if m.get("isFuture") or m.get("isCurrent")]
        for month in future_months:
            headers.append(month["monthLabel"])

        # Write headers
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            cell.font = openpyxl.styles.Font(bold=True)

        # Add example row
        sheet.cell(row=2, column=1, value="CUST-001")
        sheet.cell(row=2, column=2, value="110001")
        sheet.cell(row=2, column=3, value="TRUE")
        sheet.cell(row=2, column=4, value="")
        sheet.cell(row=2, column=5, value="Example forecast")

        # Add example quantities
        for col_idx in range(6, len(headers) + 1):
            sheet.cell(row=2, column=col_idx, value=100)

        # Auto-size columns
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width

        return workbook

    @staticmethod
    def generate_forecast_export(forecasts: List[Dict[str, Any]], cycle_months: List[Dict[str, Any]]) -> openpyxl.Workbook:
        """
        Generate an Excel export of existing forecasts
        
        Returns a workbook with all forecast data
        """
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Forecast Export"

        # Create headers
        headers = ["Customer ID", "Customer Name", "Product ID", "Product Code", "Product Description", "Status", "Submitted At"]
        
        # Add month columns
        for month in cycle_months:
            headers.append(month["monthLabel"])

        # Write headers
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            cell.font = openpyxl.styles.Font(bold=True)

        # Write forecast data
        for row_idx, forecast in enumerate(forecasts, start=2):
            # Basic forecast info
            sheet.cell(row=row_idx, column=1, value=forecast.get("customerId", ""))
            sheet.cell(row=row_idx, column=2, value=forecast.get("customerName", ""))
            sheet.cell(row=row_idx, column=3, value=forecast.get("productId", ""))
            sheet.cell(row=row_idx, column=4, value=forecast.get("productCode", ""))
            sheet.cell(row=row_idx, column=5, value=forecast.get("productDescription", ""))
            sheet.cell(row=row_idx, column=6, value=forecast.get("status", ""))
            sheet.cell(row=row_idx, column=7, value=forecast.get("submittedAt", ""))

            # Monthly forecast data
            monthly_forecasts = forecast.get("forecasts", [])
            for col_idx in range(8, len(headers) + 1):
                month_label = headers[col_idx - 1]
                # Find matching month data
                month_data = next((m for m in monthly_forecasts if m.get("monthLabel") == month_label), {})
                sheet.cell(row=row_idx, column=col_idx, value=month_data.get("quantity", 0))

        # Auto-size columns
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width

        return workbook
